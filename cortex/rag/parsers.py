"""
File format parsers for document ingestion.

Supports: PDF, DOCX, XLSX, HTML, Markdown, Plain text
"""

import logging
from typing import Optional
from pathlib import Path

# PDF parsing
try:
    from pypdf import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# DOCX parsing
try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

# Excel parsing
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# HTML parsing
try:
    from bs4 import BeautifulSoup
    HTML_AVAILABLE = True
except ImportError:
    HTML_AVAILABLE = False

# Markdown
try:
    import markdown
    MARKDOWN_AVAILABLE = True
except ImportError:
    MARKDOWN_AVAILABLE = False

logger = logging.getLogger(__name__)


class UnsupportedFileTypeError(Exception):
    """File type not supported."""
    pass


class FileParsingError(Exception):
    """Error parsing file."""
    pass


async def parse_file(
    content_bytes: bytes,
    filename: str,
    content_type: Optional[str] = None,
) -> str:
    """
    Parse file and extract text content.

    Args:
        content_bytes: Raw file bytes
        filename: Original filename
        content_type: MIME type (optional)

    Returns:
        Extracted text content

    Raises:
        UnsupportedFileTypeError: If file type not supported
        FileParsingError: If parsing fails
    """
    # Detect file type from extension
    ext = Path(filename).suffix.lower()

    # PDF
    if ext == '.pdf':
        if not PDF_AVAILABLE:
            raise UnsupportedFileTypeError("PDF support not installed (install pypdf)")
        return await parse_pdf(content_bytes)

    # DOCX
    elif ext in ['.docx', '.doc']:
        if not DOCX_AVAILABLE:
            raise UnsupportedFileTypeError("DOCX support not installed (install python-docx)")
        return await parse_docx(content_bytes)

    # Excel
    elif ext in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            raise UnsupportedFileTypeError("Excel support not installed (install openpyxl)")
        return await parse_excel(content_bytes)

    # HTML
    elif ext in ['.html', '.htm']:
        if not HTML_AVAILABLE:
            raise UnsupportedFileTypeError("HTML support not installed (install beautifulsoup4)")
        return await parse_html(content_bytes)

    # Markdown
    elif ext in ['.md', '.markdown']:
        return await parse_markdown(content_bytes)

    # Plain text
    elif ext in ['.txt', '.text', '.log']:
        return content_bytes.decode('utf-8', errors='replace')

    # Unsupported
    else:
        raise UnsupportedFileTypeError(
            f"Unsupported file type: {ext}. "
            f"Supported: .pdf, .docx, .xlsx, .html, .md, .txt"
        )


async def parse_pdf(content_bytes: bytes) -> str:
    """Extract text from PDF."""
    try:
        from io import BytesIO

        reader = PdfReader(BytesIO(content_bytes))
        text_parts = []

        for page_num, page in enumerate(reader.pages, 1):
            text = page.extract_text()
            if text.strip():
                text_parts.append(f"--- Page {page_num} ---\n{text}")

        if not text_parts:
            raise FileParsingError("PDF contains no extractable text")

        return "\n\n".join(text_parts)

    except Exception as e:
        raise FileParsingError(f"Failed to parse PDF: {str(e)}")


async def parse_docx(content_bytes: bytes) -> str:
    """Extract text from DOCX."""
    try:
        from io import BytesIO

        doc = DocxDocument(BytesIO(content_bytes))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

        if not paragraphs:
            raise FileParsingError("DOCX contains no text")

        return "\n\n".join(paragraphs)

    except Exception as e:
        raise FileParsingError(f"Failed to parse DOCX: {str(e)}")


async def parse_excel(content_bytes: bytes) -> str:
    """Extract text from Excel."""
    try:
        from io import BytesIO

        wb = load_workbook(BytesIO(content_bytes), read_only=True, data_only=True)
        text_parts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")

            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)

        if len(text_parts) <= 1:  # Only headers
            raise FileParsingError("Excel file contains no data")

        return "\n".join(text_parts)

    except Exception as e:
        raise FileParsingError(f"Failed to parse Excel: {str(e)}")


async def parse_html(content_bytes: bytes) -> str:
    """Extract text from HTML."""
    try:
        html_content = content_bytes.decode('utf-8', errors='replace')
        soup = BeautifulSoup(html_content, 'lxml')

        # Remove script and style tags
        for tag in soup(['script', 'style', 'meta', 'link']):
            tag.decompose()

        # Get text
        text = soup.get_text(separator='\n', strip=True)

        if not text.strip():
            raise FileParsingError("HTML contains no text")

        return text

    except Exception as e:
        raise FileParsingError(f"Failed to parse HTML: {str(e)}")


async def parse_markdown(content_bytes: bytes) -> str:
    """Parse Markdown (return as plain text or HTML)."""
    try:
        md_content = content_bytes.decode('utf-8', errors='replace')

        # Option 1: Return as-is (markdown text)
        return md_content

        # Option 2: Convert to HTML then extract text
        # if MARKDOWN_AVAILABLE and HTML_AVAILABLE:
        #     html = markdown.markdown(md_content)
        #     soup = BeautifulSoup(html, 'lxml')
        #     return soup.get_text(separator='\n', strip=True)
        # else:
        #     return md_content

    except Exception as e:
        raise FileParsingError(f"Failed to parse Markdown: {str(e)}")
